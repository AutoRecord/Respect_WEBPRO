#!/usr/bin/env python3
"""
WEBPRO入力シート 統合変換スクリプト
====================================

使い方:
    python convert_webpro.py <入力ディレクトリ> <出力ディレクトリ>

出力ファイル:
    - buildings.csv              建物マスタ
    - all_rooms.csv              室仕様
    - all_zones.csv              空調ゾーン
    - all_wall_specs.csv         外壁仕様（親）
    - all_wall_layers.csv        外壁層（子）
    - all_window_specs.csv       窓仕様
    - all_envelopes.csv          外皮
    - all_heat_source_groups.csv 熱源群（親）
    - all_heat_source_units.csv  熱源ユニット（子）
    - all_pump_groups.csv        二次ポンプ群（親）
    - all_pump_units.csv         二次ポンプユニット（子）
    - all_ahu_groups.csv         空調機群（親）
    - all_ahu_units.csv          空調機ユニット（子）
    - all_vent_rooms.csv         換気室
    - all_vent_fans.csv          換気送風機
    - all_lighting.csv           照明
    - all_hw_rooms.csv           給湯室
    - all_hw_equipment.csv       給湯機器
    - all_elevators.csv          昇降機
"""

import openpyxl
import csv
import os
import sys
import glob
from pathlib import Path


# ============================================================
# ユーティリティ関数
# ============================================================

def extract_priority(val):
    """運転順位を数値に変換（"1番目" → 1）"""
    if not val:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).replace('番目', '').strip()
    try:
        return int(s)
    except:
        return None


def to_bool(val):
    """■を1に変換"""
    return 1 if val == '■' else 0


def save_to_csv(data, output_path):
    """CSVに保存"""
    if not data:
        return 0
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    return len(data)


# ============================================================
# 様式0: 基本情報
# ============================================================

def extract_building_info(wb, building_id):
    ws = wb['0) 基本情報']
    info = {'building_id': building_id}
    
    for row in ws.iter_rows(min_row=1, max_row=20, max_col=6, values_only=True):
        if row[1] == '建物の名称':
            info['name'] = row[2]
        elif row[1] == '省エネ基準地域区分':
            info['region'] = row[2]
        elif row[1] == '延べ面積 [㎡]':
            info['total_area'] = row[2]
        elif row[1] == '年間日射地域区分':
            info['solar_region'] = row[2]
        elif row[1] and '都道府県' in str(row[2]):
            info['prefecture'] = row[3]
    
    return info


# ============================================================
# 様式1: 室仕様
# ============================================================

def convert_rooms(wb, building_id):
    ws = wb['1) 室仕様']
    rooms = []
    counter = 1
    
    for row in ws.iter_rows(min_row=11, max_col=14, values_only=True):
        if not row[0] and not row[1]:
            continue
        
        rooms.append({
            'building_id': building_id,
            'room_id': f'{building_id}_R{counter:03d}',
            'floor': row[0],
            'room_name': row[1],
            'building_use': row[2],
            'room_use_major': row[3],
            'room_use_minor': row[4],
            'area': row[5],
            'floor_height': row[6],
            'ceiling_height': row[7],
            'is_ac_target': to_bool(row[8]),
            'is_vent_target': to_bool(row[9]),
            'is_light_target': to_bool(row[10]),
            'is_hw_target': to_bool(row[11]),
            'sub_building_name': row[12] if row[12] else '',
            'note': row[13] if row[13] else ''
        })
        counter += 1
    
    return rooms


# ============================================================
# 様式2-1: 空調ゾーン
# ============================================================

def convert_zones(wb, building_id):
    ws = wb['2-1) 空調ゾーン']
    zones = []
    counter = 1
    
    for row in ws.iter_rows(min_row=11, max_col=12, values_only=True):
        if not row[0] and not row[1]:
            continue
        
        zones.append({
            'building_id': building_id,
            'zone_id': f'{building_id}_Z{counter:03d}',
            'room_floor': row[0],
            'room_name': row[1],
            'room_use_major': row[2],
            'room_use_minor': row[3],
            'room_area': row[4],
            'floor_height': row[5],
            'ceiling_height': row[6],
            'zone_floor': row[7],
            'zone_name': row[8],
            'ahu_group_room_load': row[9],
            'ahu_group_oa_load': row[10],
            'note': row[11] if row[11] else ''
        })
        counter += 1
    
    return zones


# ============================================================
# 様式2-2: 外壁構成（親子分解）
# ============================================================

def convert_walls(wb, building_id):
    ws = wb['2-2) 外壁構成 ']
    specs = []
    layers = []
    
    current_wall = None
    layer_order = 0
    
    for row in ws.iter_rows(min_row=11, max_col=9, values_only=True):
        wall_name, wall_type, u_value, material_no, material_name, conductivity, thickness, absorption, note = row[:9]
        
        if wall_name:
            if current_wall:
                specs.append(current_wall)
            current_wall = {
                'building_id': building_id,
                'wall_spec_id': f'{building_id}_{wall_name}',
                'wall_name': wall_name,
                'wall_type': wall_type,
                'u_value': u_value,
                'absorption': absorption,
                'note': note if note else ''
            }
            layer_order = 0
            continue
        
        if material_name in ['室内側', '室外側']:
            continue
        
        if material_no and current_wall:
            layer_order += 1
            layers.append({
                'building_id': building_id,
                'wall_spec_id': current_wall['wall_spec_id'],
                'layer_order': layer_order,
                'material_no': material_no,
                'material_name': material_name,
                'conductivity': conductivity,
                'thickness_mm': thickness
            })
    
    if current_wall:
        specs.append(current_wall)
    
    return specs, layers


# ============================================================
# 様式2-3: 窓仕様
# ============================================================

def convert_windows(wb, building_id):
    ws = wb['2-3) 窓仕様']
    windows = []
    
    for row in ws.iter_rows(min_row=11, max_col=8, values_only=True):
        if not row[0]:
            continue
        
        windows.append({
            'building_id': building_id,
            'window_spec_id': f'{building_id}_{row[0]}',
            'window_name': row[0],
            'u_value_total': row[1],
            'eta_value_total': row[2],
            'frame_type': row[3],
            'glass_type': row[4],
            'u_value_glass': row[5],
            'eta_value_glass': row[6],
            'note': row[7] if row[7] else ''
        })
    
    return windows


# ============================================================
# 様式2-4: 外皮（親子分解）
# ============================================================

def convert_envelopes(wb, building_id):
    ws = wb['2-4) 外皮 ']
    envelopes = []
    counter = 1
    current_floor = None
    current_zone = None
    
    for row in ws.iter_rows(min_row=11, max_col=11, values_only=True):
        floor, zone_name, direction = row[0], row[1], row[2]
        
        if not direction and not row[5]:
            continue
        
        if floor and zone_name:
            current_floor = floor
            current_zone = zone_name
        
        if current_floor and current_zone:
            envelopes.append({
                'building_id': building_id,
                'envelope_id': f'{building_id}_E{counter:03d}',
                'floor': current_floor,
                'zone_name': current_zone,
                'direction': direction,
                'shade_coef_cool': row[3],
                'shade_coef_heat': row[4],
                'wall_spec_name': row[5],
                'wall_area': row[6],
                'window_spec_name': row[7] if row[7] else None,
                'window_area': row[8] if row[8] else None,
                'blind': row[9] if row[9] else None,
                'note': row[10] if row[10] else ''
            })
            counter += 1
    
    return envelopes


# ============================================================
# 様式2-5: 熱源（親子分解）
# ============================================================

def convert_heat_sources(wb, building_id):
    ws = wb['2-5) 熱源']
    groups = []
    units = []
    
    current_group = None
    unit_counter = 0
    
    for row in ws.iter_rows(min_row=11, max_col=24, values_only=True):
        group_name = row[0]
        equipment_type = row[5]
        
        if not equipment_type and not group_name:
            continue
        
        if group_name:
            current_group = {
                'building_id': building_id,
                'hs_group_id': f'{building_id}_{group_name}',
                'group_name': group_name,
                'simultaneous_supply': row[1],
                'unit_control': row[2],
                'operation_mode': row[3],
                'storage_capacity_mj': row[4] if row[4] else None,
                'note': row[23] if len(row) > 23 and row[23] else ''
            }
            groups.append(current_group)
            unit_counter = 0
        
        if equipment_type and current_group:
            unit_counter += 1
            units.append({
                'building_id': building_id,
                'hs_unit_id': f"{current_group['hs_group_id']}_U{unit_counter:02d}",
                'hs_group_id': current_group['hs_group_id'],
                'equipment_type': equipment_type,
                'priority_cool': extract_priority(row[6]),
                'quantity_cool': row[7],
                'supply_temp_cool': row[8],
                'cooling_capacity_kw': row[9],
                'main_power_cool_kw': row[10],
                'aux_power_cool_kw': row[11],
                'pump_power_cool_kw': row[12],
                'ct_capacity_kw': row[13],
                'ct_fan_power_kw': row[14],
                'ct_pump_power_kw': row[15],
                'priority_heat': extract_priority(row[16]),
                'quantity_heat': row[17],
                'supply_temp_heat': row[18],
                'heating_capacity_kw': row[19],
                'main_power_heat_kw': row[20] if len(row) > 20 else None,
                'aux_power_heat_kw': row[21] if len(row) > 21 else None,
                'pump_power_heat_kw': row[22] if len(row) > 22 else None,
            })
    
    return groups, units


# ============================================================
# 様式2-6: 二次ポンプ（親子分解）
# ============================================================

def convert_pumps(wb, building_id):
    ws = wb['2-6) 2次ﾎﾟﾝﾌﾟ']
    groups = []
    units = []
    
    current_group = None
    unit_counter = 0
    
    for row in ws.iter_rows(min_row=11, max_col=11, values_only=True):
        group_name, unit_control, temp_cool, temp_heat, priority = row[:5]
        
        if not priority and not group_name:
            continue
        
        if group_name:
            current_group = {
                'building_id': building_id,
                'pump_group_id': f'{building_id}_{group_name}',
                'group_name': group_name,
                'unit_control': unit_control,
                'temp_diff_cool': temp_cool,
                'temp_diff_heat': temp_heat,
                'note': row[10] if row[10] else ''
            }
            groups.append(current_group)
            unit_counter = 0
        
        if priority and current_group:
            unit_counter += 1
            units.append({
                'building_id': building_id,
                'pump_unit_id': f"{current_group['pump_group_id']}_U{unit_counter:02d}",
                'pump_group_id': current_group['pump_group_id'],
                'priority': extract_priority(priority),
                'quantity': row[5],
                'flow_rate_m3h': row[6],
                'power_kw': row[7],
                'flow_control': row[8],
                'min_flow_ratio_pct': row[9]
            })
    
    return groups, units


# ============================================================
# 様式2-7: 空調機（親子分解）
# ============================================================

def convert_ahus(wb, building_id):
    ws = wb['2-7) 空調機']
    groups = []
    units = []
    
    current_group = None
    unit_counter = 0
    
    for row in ws.iter_rows(min_row=11, max_col=26, values_only=True):
        group_name = row[0]
        ahu_type = row[2]
        
        if not ahu_type and not group_name:
            continue
        
        if group_name:
            current_group = {
                'building_id': building_id,
                'ahu_group_id': f'{building_id}_{group_name}',
                'group_name': group_name,
                'pump_group_cool': row[19] if len(row) > 19 else None,
                'pump_group_heat': row[20] if len(row) > 20 else None,
                'hs_group_cool': row[21] if len(row) > 21 else None,
                'hs_group_heat': row[22] if len(row) > 22 else None,
                'note': row[23] if len(row) > 23 and row[23] else ''
            }
            groups.append(current_group)
            unit_counter = 0
        
        if ahu_type and current_group:
            unit_counter += 1
            units.append({
                'building_id': building_id,
                'ahu_unit_id': f"{current_group['ahu_group_id']}_U{unit_counter:02d}",
                'ahu_group_id': current_group['ahu_group_id'],
                'quantity': row[1],
                'ahu_type': ahu_type,
                'cooling_capacity_kw': row[3],
                'heating_capacity_kw': row[4],
                'design_oa_flow_m3h': row[5],
                'fan_supply_kw': row[6],
                'fan_return_kw': row[7],
                'fan_oa_kw': row[8],
                'fan_exhaust_kw': row[9],
                'volume_control': row[10],
                'min_volume_ratio_pct': row[11],
                'preheat_oa_stop': row[12],
                'oa_cooling_control': row[13],
                'hex_type': row[14]
            })
    
    return groups, units


# ============================================================
# 様式3: 換気
# ============================================================

def convert_vent_rooms(wb, building_id):
    ws = wb['3-1) 換気室']
    records = []
    counter = 1
    
    for row in ws.iter_rows(min_row=11, max_col=8, values_only=True):
        if not row[0] and not row[1] and not row[5]:
            continue
        
        records.append({
            'building_id': building_id,
            'vent_room_id': f'{building_id}_VR{counter:03d}',
            'floor': row[0],
            'room_name': row[1],
            'room_use_major': row[2],
            'room_use_minor': row[3],
            'area': row[4],
            'vent_type': row[5],
            'equipment_name': row[6],
            'note': row[7] if row[7] else ''
        })
        counter += 1
    
    return records


def convert_vent_fans(wb, building_id):
    ws = wb['3-2) 換気送風機']
    records = []
    counter = 1
    
    for row in ws.iter_rows(min_row=11, max_col=7, values_only=True):
        if not row[0]:
            continue
        
        records.append({
            'building_id': building_id,
            'vent_fan_id': f'{building_id}_VF{counter:03d}',
            'equipment_name': row[0],
            'design_flow_m3h': row[1],
            'motor_power_kw': row[2],
            'high_efficiency': row[3],
            'inverter': row[4],
            'flow_control': row[5],
            'note': row[6] if row[6] else ''
        })
        counter += 1
    
    return records


# ============================================================
# 様式4: 照明（親子分解）
# ============================================================

def convert_lighting(wb, building_id):
    ws = wb['4) 照明']
    records = []
    counter = 1
    current_room = None
    
    for row in ws.iter_rows(min_row=11, max_col=18, values_only=True):
        floor, room_name = row[0], row[1]
        fixture_name = row[10]
        
        if not fixture_name:
            continue
        
        if floor and room_name:
            current_room = {
                'floor': floor, 'room_name': room_name,
                'use_major': row[2], 'use_minor': row[3],
                'area': row[4], 'ceiling_height': row[6]
            }
        
        if current_room:
            records.append({
                'building_id': building_id,
                'lighting_id': f'{building_id}_LT{counter:03d}',
                'floor': current_room['floor'],
                'room_name': current_room['room_name'],
                'room_use_major': current_room['use_major'],
                'room_use_minor': current_room['use_minor'],
                'area': current_room['area'],
                'ceiling_height': current_room['ceiling_height'],
                'fixture_name': fixture_name,
                'power_per_unit_w': row[11],
                'quantity': row[12],
                'occupancy_control': row[13] if row[13] else None,
                'daylight_control': row[14] if row[14] else None,
                'schedule_control': row[15] if row[15] else None,
                'initial_lumen_correction': row[16] if row[16] else None,
                'note': row[17] if row[17] else ''
            })
            counter += 1
    
    return records


# ============================================================
# 様式5: 給湯
# ============================================================

def convert_hw_rooms(wb, building_id):
    ws = wb['5-1) 給湯室']
    records = []
    counter = 1
    current_room = None
    
    for row in ws.iter_rows(min_row=11, max_col=9, values_only=True):
        floor, room_name, tap_location = row[0], row[1], row[5]
        
        if not tap_location and not floor:
            continue
        
        if floor and room_name:
            current_room = {
                'floor': floor, 'room_name': room_name,
                'use_major': row[2], 'use_minor': row[3], 'area': row[4]
            }
        
        if current_room and tap_location:
            records.append({
                'building_id': building_id,
                'hw_room_id': f'{building_id}_HW{counter:03d}',
                'floor': current_room['floor'],
                'room_name': current_room['room_name'],
                'room_use_major': current_room['use_major'],
                'room_use_minor': current_room['use_minor'],
                'area': current_room['area'],
                'tap_location': tap_location,
                'water_saving_device': row[6],
                'equipment_name': row[7],
                'note': row[8] if row[8] else ''
            })
            counter += 1
    
    return records


def convert_hw_equipment(wb, building_id):
    ws = wb['5-2) 給湯機器']
    records = []
    counter = 1
    
    for row in ws.iter_rows(min_row=11, max_col=10, values_only=True):
        if not row[0]:
            continue
        
        records.append({
            'building_id': building_id,
            'hw_equip_id': f'{building_id}_HE{counter:03d}',
            'equipment_name': row[0],
            'fuel_type': row[1],
            'heating_capacity_kw': row[2],
            'efficiency': row[3],
            'pipe_insulation': row[4],
            'pipe_size_mm': row[5],
            'solar_collector_area_m2': row[6] if row[6] else None,
            'collector_azimuth_deg': row[7] if row[7] else None,
            'collector_tilt_deg': row[8] if row[8] else None,
            'note': row[9] if row[9] else ''
        })
        counter += 1
    
    return records


# ============================================================
# 様式6: 昇降機
# ============================================================

def convert_elevators(wb, building_id):
    ws = wb['6) 昇降機']
    records = []
    counter = 1
    
    for row in ws.iter_rows(min_row=11, max_col=11, values_only=True):
        if not row[4]:
            continue
        
        records.append({
            'building_id': building_id,
            'elevator_id': f'{building_id}_EV{counter:03d}',
            'floor': row[0],
            'room_name': row[1],
            'room_use_major': row[2],
            'room_use_minor': row[3],
            'equipment_name': row[4],
            'quantity': row[5],
            'load_capacity_kg': row[6],
            'speed_m_min': row[7],
            'transport_coefficient': row[8],
            'speed_control': row[9],
            'note': row[10] if row[10] else ''
        })
        counter += 1
    
    return records


# ============================================================
# 様式7-1: 太陽光発電
# ============================================================

def convert_pv(wb, building_id):
    ws = wb['7-1) 太陽光発電']
    records = []
    counter = 1
    
    for row in ws.iter_rows(min_row=11, max_col=8, values_only=True):
        if not row[0]:
            continue
        
        records.append({
            'building_id': building_id,
            'pv_id': f'{building_id}_PV{counter:03d}',
            'system_name': row[0],
            'pcs_efficiency': row[1],
            'cell_type': row[2],
            'mount_type': row[3],
            'capacity_kw': row[4],
            'azimuth_deg': row[5],
            'tilt_deg': row[6],
            'note': row[7] if row[7] else ''
        })
        counter += 1
    
    return records


# ============================================================
# 様式7-3: コージェネレーション
# ============================================================

def convert_cogen(wb, building_id):
    ws = wb['7-3) コージェネレーション設備']
    records = []
    counter = 1
    
    for row in ws.iter_rows(min_row=11, max_col=12, values_only=True):
        if not row[0]:
            continue
        
        records.append({
            'building_id': building_id,
            'cogen_id': f'{building_id}_CG{counter:03d}',
            'equipment_name': row[0],
            'rated_power_kw': row[1],
            'quantity': row[2],
            'power_eff_100': row[3],
            'power_eff_75': row[4],
            'power_eff_50': row[5],
            'heat_eff_100': row[6],
            'heat_eff_75': row[7],
            'heat_eff_50': row[8],
            'ac_cool_priority': extract_priority(row[9]),
            'ac_heat_priority': extract_priority(row[10]),
            'hw_priority': extract_priority(row[11])
        })
        counter += 1
    
    return records


# ============================================================
# 様式8: 非空調外皮（親子分解）
# ============================================================

def convert_non_ac_envelope(wb, building_id):
    ws = wb['8) 非空調外皮']
    records = []
    counter = 1
    current_zone = None
    
    for row in ws.iter_rows(min_row=11, max_col=14, values_only=True):
        floor = row[0]
        zone_name = row[1]
        direction = row[6]
        
        if not direction and not floor:
            continue
        
        if floor and zone_name:
            current_zone = {
                'floor': floor,
                'zone_name': zone_name,
                'use_major': row[2],
                'use_minor': row[3],
                'area': row[4],
                'floor_height': row[5]
            }
        
        if current_zone and direction:
            records.append({
                'building_id': building_id,
                'non_ac_envelope_id': f'{building_id}_NAE{counter:03d}',
                'floor': current_zone['floor'],
                'zone_name': current_zone['zone_name'],
                'room_use_major': current_zone['use_major'],
                'room_use_minor': current_zone['use_minor'],
                'area': current_zone['area'],
                'floor_height': current_zone['floor_height'],
                'direction': direction,
                'shade_coef_cool': row[7],
                'shade_coef_heat': row[8],
                'wall_spec_name': row[9],
                'wall_area': row[10],
                'window_spec_name': row[11] if row[11] else None,
                'window_area': row[12] if len(row) > 12 and row[12] else None,
                'note': row[13] if len(row) > 13 and row[13] else ''
            })
            counter += 1
    
    return records


# ============================================================
# メイン処理
# ============================================================

def process_file(filepath, building_id):
    """1ファイルを処理"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    result = {
        'building': extract_building_info(wb, building_id),
        'rooms': convert_rooms(wb, building_id),
        'zones': convert_zones(wb, building_id),
    }
    
    result['wall_specs'], result['wall_layers'] = convert_walls(wb, building_id)
    result['window_specs'] = convert_windows(wb, building_id)
    result['envelopes'] = convert_envelopes(wb, building_id)
    result['hs_groups'], result['hs_units'] = convert_heat_sources(wb, building_id)
    result['pump_groups'], result['pump_units'] = convert_pumps(wb, building_id)
    result['ahu_groups'], result['ahu_units'] = convert_ahus(wb, building_id)
    result['vent_rooms'] = convert_vent_rooms(wb, building_id)
    result['vent_fans'] = convert_vent_fans(wb, building_id)
    result['lighting'] = convert_lighting(wb, building_id)
    result['hw_rooms'] = convert_hw_rooms(wb, building_id)
    result['hw_equipment'] = convert_hw_equipment(wb, building_id)
    result['elevators'] = convert_elevators(wb, building_id)
    result['pv'] = convert_pv(wb, building_id)
    result['cogen'] = convert_cogen(wb, building_id)
    result['non_ac_envelope'] = convert_non_ac_envelope(wb, building_id)
    
    return result


def main(input_dir, output_dir):
    """メイン処理"""
    os.makedirs(output_dir, exist_ok=True)
    
    # 入力ファイルを検索
    patterns = ['*WEBPRO*.xlsx', '*webpro*.xlsx', '*.xlsx']
    files = []
    for pattern in patterns:
        files.extend(glob.glob(os.path.join(input_dir, pattern)))
    files = list(set(files))  # 重複除去
    
    if not files:
        print(f'Error: No Excel files found in {input_dir}')
        return
    
    print(f'Found {len(files)} files')
    
    # 集約用
    all_data = {
        'buildings': [],
        'rooms': [], 'zones': [],
        'wall_specs': [], 'wall_layers': [],
        'window_specs': [], 'envelopes': [],
        'hs_groups': [], 'hs_units': [],
        'pump_groups': [], 'pump_units': [],
        'ahu_groups': [], 'ahu_units': [],
        'vent_rooms': [], 'vent_fans': [],
        'lighting': [],
        'hw_rooms': [], 'hw_equipment': [],
        'elevators': [],
        'pv': [], 'cogen': [], 'non_ac_envelope': []
    }
    
    for filepath in sorted(files):
        filename = os.path.basename(filepath)
        # building_idをファイル名から生成
        building_id = Path(filepath).stem
        for suffix in ['_WEBPRO_inputSheet_for_Ver3_9', '_webpro', '_WEBPRO']:
            building_id = building_id.replace(suffix, '')
        
        print(f'\nProcessing: {filename} -> {building_id}')
        
        try:
            result = process_file(filepath, building_id)
            
            all_data['buildings'].append(result['building'])
            for key in all_data.keys():
                if key != 'buildings' and key in result:
                    all_data[key].extend(result[key])
            
            print(f'  Rooms: {len(result["rooms"])}, Zones: {len(result["zones"])}')
            print(f'  HS: {len(result["hs_groups"])} groups, AHU: {len(result["ahu_groups"])} groups')
            print(f'  Lighting: {len(result["lighting"])}, Elevators: {len(result["elevators"])}')
            
        except Exception as e:
            print(f'  Error: {e}')
    
    # CSV出力
    print(f'\n=== Saving to {output_dir} ===')
    
    output_map = {
        'buildings': 'buildings.csv',
        'rooms': 'all_rooms.csv',
        'zones': 'all_zones.csv',
        'wall_specs': 'all_wall_specs.csv',
        'wall_layers': 'all_wall_layers.csv',
        'window_specs': 'all_window_specs.csv',
        'envelopes': 'all_envelopes.csv',
        'hs_groups': 'all_heat_source_groups.csv',
        'hs_units': 'all_heat_source_units.csv',
        'pump_groups': 'all_pump_groups.csv',
        'pump_units': 'all_pump_units.csv',
        'ahu_groups': 'all_ahu_groups.csv',
        'ahu_units': 'all_ahu_units.csv',
        'vent_rooms': 'all_vent_rooms.csv',
        'vent_fans': 'all_vent_fans.csv',
        'lighting': 'all_lighting.csv',
        'hw_rooms': 'all_hw_rooms.csv',
        'hw_equipment': 'all_hw_equipment.csv',
        'elevators': 'all_elevators.csv',
        'pv': 'all_pv.csv',
        'cogen': 'all_cogen.csv',
        'non_ac_envelope': 'all_non_ac_envelope.csv'
    }
    
    for key, filename in output_map.items():
        path = os.path.join(output_dir, filename)
        count = save_to_csv(all_data[key], path)
        print(f'  {filename}: {count} records')
    
    print('\n✅ Done!')


if __name__ == '__main__':
    if len(sys.argv) >= 3:
        input_dir = sys.argv[1]
        output_dir = sys.argv[2]
    else:
        input_dir = '/mnt/user-data/uploads'
        output_dir = '/home/claude/output'
    
    main(input_dir, output_dir)
